from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import json

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9030")
driver = webdriver.Chrome(service=service, options=options)
driver.get('https://www.magazineluiza.com.br/artesanato/l/am/')

output = []
total_company = 0
dis = 500
for price in range(87500, 100000, dis):
    try:
        min_price_input = Find_Element(driver, By.XPATH, '//*[@id="__next"]/div/main/section[3]/div[2]/div[3]/div[2]/div/div[1]/div[1]/input')
        max_price_input = Find_Element(driver, By.XPATH, '//*[@id="__next"]/div/main/section[3]/div[2]/div[3]/div[2]/div/div[1]/div[2]/input')
        min_price = Send_Keys(min_price_input, f'{price}')
        max_price = Send_Keys(max_price_input, f'{price + dis}')
        button = Find_Element(driver, By.XPATH, '//*[@id="__next"]/div/main/section[3]/div[2]/div[3]/div[2]/div/button')
        driver.execute_script("arguments[0].click();", button)
        print('done')
        categories = Find_Elements(driver, By.XPATH, '//*[@id="__next"]/div/main/section[3]/div[3]/div')
        for category in categories:
            category_names = category.text.split('\n')
            if category_names[0] == 'Vendido por':
                try:
                    add_button = category.find_element(By.TAG_NAME, 'button')
                    driver.execute_script("arguments[0].click();", add_button)
                except:
                    pass
                company_names = category.find_elements(By.TAG_NAME, 'li')
                print(f'{price} ~ {price + dis} : company_number --> {len(company_names)}')
                cur_num = len(company_names)
                total_company += cur_num
                print(total_company)
                for company_name in company_names:
                    print(company_name.text)
                    output.append({'company_name' : company_name.text})
                break
        with open('output(10^4-10^5).json', 'w') as file:
            json.dump(output, file)
        print('done')
        driver.back()
        sleep(1)
    except:
        pass






                # for sold in solds:
                #     company_name = sold.text
                #     output.append({'company_name' : company_name})
                #     if company_name == 'Magalu':
                #         output.remove('Magalu')

                # for sold_index in range(len(solds)):
                    # driver.get(f'https://www.magazineluiza.com.br/celulares-e-smartphones/l/te/price---{price - 3000}:{price}+review---{star}+seller---capas25/?sfilters=0')


                # products = Find_Element(driver, By.XPATH, '//*[@id="__next"]/div/main/section[4]/div[1]/div/p').text.split(' ')
                # print(f'numbers --> {products[0]}')

            # if category_name == 'Vendido por':
            #     solds = Find_Elements(driver, By.XPATH, f'//*[@id="__next"]/div/main/section[3]/div[3]/div[{i}]/div[2]/ul/li')
            #     print(f'category[{price-5000} : {price}, {star}] --> {len(solds)}')







# wb = Workbook()
# sheet = wb.active

# border_style = Side(border_style = "thin", color = "000000")
# font = Font(name = 'Times New Roman', size = '12')
# alignment = Alignment(vertical = 'center', horizontal = 'center')

# item = ["Continent", "Country", "River", "URL"]

# for i in range(0, 4):
#     sheet.cell(row = 1, column = i + 1).value = item[i]
#     sheet.cell(row = 1, column = i + 1).border = Border(right = border_style, bottom = border_style)
#     sheet.cell(row = 1, column = i + 1).font = font
#     sheet.cell(row = 1, column = i + 1).alignment = alignment

# output = []

# start_row = 2
# for id in range(74, 81):
#     continents = Find_Elements(driver, By.ID, f'ref3270{id}')
#     for continent in continents:
#         continent_name = continent.find_element(By.TAG_NAME, 'a').text
#         print(continent_name)
#         sheet.cell(row = start_row, column = 1).value = continent_name
#         countries = continent.find_elements(By.XPATH, f'//*[@id="ref3270{id}"]/ul/li')
#         print(len(countries))
#         for country_index in range(len(countries)):
#             country_name = countries[country_index].find_element(By.TAG_NAME, 'a').text
#             print(country_name)
#             sheet.cell(row = start_row, column = 2).value = country_name
#             rivers = countries[country_index].find_elements(By.XPATH, f'//*[@id="ref3270{id}"]/ul/li[{country_index + 1}]/div/ul/li')
#             print(len(rivers))
#             for river in rivers:
#                 try:
#                     river_name = river.find_element(By.TAG_NAME, 'div').text
#                     print(river_name)
#                     sheet.cell(row = start_row, column = 3).value = river_name
#                     river_url = river.find_element(By.TAG_NAME, 'a').get_attribute('href')
#                     sheet.cell(row = start_row, column = 4).value = river_url
#                     output.append({"River" : river_name, "URL" : river_url})
#                 except:
#                     pass
#                 start_row += 1

# with open('output.json', 'w') as file:
#     json.dump(output, file)

# wb.save('output.xlsx')